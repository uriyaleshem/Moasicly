from __future__ import annotations

import argparse
import csv
import json
import tempfile
import time
import zipfile
from pathlib import Path
import sys
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from class_balancer.db import Database
from class_balancer.services import AssignmentService, ExportService, ImportService, ProjectService, ReportService


STUDENT_COUNT = 230
CLASS_COUNT = 8
HEADERS = [
    "internal_code",
    "full_name",
    "gender",
    "source_school",
    "math_grade",
    "english_grade",
    "hebrew_grade",
    "average",
    "behavior",
    "dominance",
    "friend 1",
    "friend 2",
    "friend 3",
    "must be with",
    "must not be with",
]


def main() -> None:
    parser = argparse.ArgumentParser(description="Run the 230-student demo verification scenario.")
    parser.add_argument("--keep-files", action="store_true", help="Keep the generated database, CSV, and XLSX export.")
    parser.add_argument("--variants", type=int, default=2, help="Assignment variants to check for each run.")
    args = parser.parse_args()

    if args.keep_files:
        root = Path("tmp_profile") / f"demo_230_{int(time.time())}"
        root.mkdir(parents=True, exist_ok=True)
        cleanup = None
    else:
        cleanup = tempfile.TemporaryDirectory()
        root = Path(cleanup.name)

    try:
        csv_path = root / "demo_230_students.csv"
        db_path = root / "demo_230.sqlite3"
        export_path = root / "demo_230_export.xlsx"
        _write_demo_csv(csv_path)

        database = Database(db_path)
        project_service = ProjectService(database)
        import_service = ImportService(database)
        assignment_service = AssignmentService(database)
        export_service = ExportService(database)
        report_service = ReportService(database)

        project_id = project_service.create_project(
            name="demo 230",
            grade_level="7",
            school_year="2026",
            class_count=CLASS_COUNT,
            class_names_text=", ".join(f"Class {index}" for index in range(1, CLASS_COUNT + 1)),
        )
        _set_class_targets(database, project_id)
        _save_settings(project_service, project_id, friendship_first=False)

        import_service.load_preview(csv_path)
        import_result = import_service.save_imported_students(project_id)

        without_priority, without_progress, without_elapsed = _run_assignment(
            assignment_service,
            project_service,
            project_id,
            friendship_first=False,
            variants=args.variants,
        )
        with_priority, with_progress, with_elapsed = _run_assignment(
            assignment_service,
            project_service,
            project_id,
            friendship_first=True,
            variants=args.variants,
        )

        dashboard = assignment_service.dashboard(project_id)
        report = report_service.quality_report(project_id)
        exported = export_service.export_project(project_id, export_path)
        export_rows = _exported_student_count(exported)
        active_before_close = database.get_active_assignment_version(project_id)
        database.close()

        reopened = Database(db_path)
        try:
            reopened_service = AssignmentService(reopened)
            reloaded_dashboard = reopened_service.dashboard(project_id)
        finally:
            reopened.close()

        with_metrics = _metrics(with_priority)
        without_metrics = _metrics(without_priority)
        checks = _checks(
            dashboard=dashboard,
            reloaded_dashboard=reloaded_dashboard,
            report=report,
            active_before_close=active_before_close,
            export_rows=export_rows,
            expected_students=STUDENT_COUNT,
        )
        result = {
            "paths": {
                "root": str(root),
                "csv": str(csv_path),
                "database": str(db_path),
                "export": str(exported),
                "kept": bool(args.keep_files),
            },
            "import": import_result,
            "without_priority": without_metrics,
            "with_priority": with_metrics,
            "improvement": {
                "students_with_friend_delta": with_metrics["students_with_friend"]
                - without_metrics["students_with_friend"],
                "fulfilled_request_delta": with_metrics["fulfilled_friend_requests"]
                - without_metrics["fulfilled_friend_requests"],
                "score_delta": round(with_metrics["score"] - without_metrics["score"], 2),
            },
            "runtime_seconds": {
                "without_priority": round(without_elapsed, 2),
                "with_priority": round(with_elapsed, 2),
                "total_assignment": round(without_elapsed + with_elapsed, 2),
            },
            "progress": {
                "without_priority_samples": len(without_progress),
                "with_priority_samples": len(with_progress),
                "without_priority_last": without_progress[-1] if without_progress else None,
                "with_priority_last": with_progress[-1] if with_progress else None,
            },
            "checks": checks,
        }
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
        if cleanup is None:
            print(f"Kept verification files in {root}")
    finally:
        if cleanup is not None:
            cleanup.cleanup()


def _write_demo_csv(path: Path) -> None:
    rows = [_student_row(index) for index in range(1, STUDENT_COUNT + 1)]
    by_index = {index: row for index, row in enumerate(rows, start=1)}

    for index in range(1, 201, 2):
        by_index[index]["friend 1"] = _student_name(index + 1)
        by_index[index + 1]["friend 1"] = _student_name(index)

    for index in range(201, STUDENT_COUNT + 1):
        by_index[index]["friend 1"] = _student_name(196 + ((index - 201) % 5))

    for index in range(1, 25, 6):
        by_index[index]["friend 2"] = _student_name(index + 5)
        by_index[index]["must not be with"] = _student_name(index + 5)

    for index in range(41, 81, 10):
        by_index[index]["must be with"] = _student_name(index + 2)
        by_index[index + 2]["must not be with"] = _student_name(index + 3)

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def _student_row(index: int) -> dict[str, Any]:
    base_grade = [96, 93, 90, 86, 82, 78, 74, 70, 66, 62][index % 10] + ((index * 7) % 5) - 2
    return {
        "internal_code": f"S{index:03d}",
        "full_name": _student_name(index),
        "gender": "M" if index % 2 else "F",
        "source_school": f"School {((index - 1) % 12) + 1}",
        "math_grade": base_grade + ((index % 3) - 1),
        "english_grade": base_grade + (((index + 1) % 3) - 1),
        "hebrew_grade": base_grade + (((index + 2) % 3) - 1),
        "average": base_grade,
        "behavior": ["high", "medium", "low"][index % 3],
        "dominance": (index * 11) % 100,
        "friend 1": "",
        "friend 2": "",
        "friend 3": "",
        "must be with": "",
        "must not be with": "",
    }


def _student_name(index: int) -> str:
    return f"Student {index:03d}"


def _set_class_targets(database: Database, project_id: int) -> None:
    for group in database.get_classes(project_id):
        database.connection.execute(
            "UPDATE classes SET min_students = ?, max_students = ?, target_students = ? WHERE id = ?",
            (28, 29, 29, int(group.id)),
        )
    database.connection.commit()


def _save_settings(project_service: ProjectService, project_id: int, friendship_first: bool) -> None:
    project_service.update_settings(
        project_id,
        {
            "friendship": True,
            "friendship_first": friendship_first,
            "friendship_weight": 0.55,
            "friendship_priority_order": False,
            "balance_gender": True,
            "balance_grades": True,
            "balance_behavior": True,
            "spread_source_school": True,
            "spread_dominant_students": True,
            "hard_class_capacity": True,
            "class_size_weight": 4.0,
            "gender_weight": 1.0,
            "grade_weight": 1.1,
            "behavior_weight": 1.0,
            "source_school_weight": 0.75,
            "dominance_weight": 0.5,
            "optimizer_backend": "local",
            "search_restarts": 5,
            "max_iterations": 120,
            "stop_when_score_at_least": 92,
            "random_seed": 1039,
            "ai_assisted_assignment": False,
            "ai_auto_review": False,
        },
    )


def _run_assignment(
    assignment_service: AssignmentService,
    project_service: ProjectService,
    project_id: int,
    *,
    friendship_first: bool,
    variants: int,
) -> tuple[dict[str, Any], list[tuple[float, str]], float]:
    _save_settings(project_service, project_id, friendship_first=friendship_first)
    progress: list[tuple[float, str]] = []

    def on_progress(percent: float, message: str) -> None:
        progress.append((round(float(percent), 2), str(message)))

    start = time.perf_counter()
    result = assignment_service.run_assignment(project_id, variant_count=variants, progress_callback=on_progress)
    elapsed = time.perf_counter() - start
    return result, progress, elapsed


def _metrics(result: dict[str, Any]) -> dict[str, Any]:
    score = result["score"]
    friendship = score.get("friendship", {}) or {}
    assignments = result["assignments"]
    class_sizes = [int(item.get("size", 0) or 0) for item in score.get("class_stats", [])]
    return {
        "score": float(score.get("total_score", 0) or 0),
        "assigned_students": len(assignments),
        "unique_assigned_students": len(set(assignments)),
        "hard_violations": len(score.get("hard_violations", []) or []),
        "students_with_friend": len(friendship.get("satisfied", []) or []),
        "students_without_friend": len(friendship.get("missing", []) or []),
        "fulfilled_friend_requests": int(friendship.get("total_received", 0) or 0),
        "total_friend_requests": int(friendship.get("total_requested", 0) or 0),
        "class_sizes": class_sizes,
        "friendship_first_active": bool(score.get("friendship_first_active", False)),
        "candidate_note": score.get("candidate_note", {}),
        "local_search_telemetry": score.get("local_search_telemetry", {}),
    }


def _checks(
    *,
    dashboard: dict[str, Any],
    reloaded_dashboard: dict[str, Any],
    report: dict[str, Any],
    active_before_close: dict[str, Any] | None,
    export_rows: int,
    expected_students: int,
) -> dict[str, Any]:
    rows = dashboard.get("rows", []) or []
    student_ids = [int(row["student_id"]) for row in rows]
    class_sizes = [int(item.get("size", 0) or 0) for item in dashboard.get("score", {}).get("class_stats", [])]
    reloaded_rows = reloaded_dashboard.get("rows", []) or []
    return {
        "dashboard_has_assignment": bool(dashboard.get("has_assignment")),
        "report_has_assignment": bool(report.get("has_assignment")),
        "active_version_saved": active_before_close is not None,
        "all_students_assigned_once": len(student_ids) == expected_students and len(set(student_ids)) == expected_students,
        "reload_preserved_rows": len(reloaded_rows) == expected_students,
        "reload_preserved_score": reloaded_dashboard.get("score", {}).get("total_score")
        == dashboard.get("score", {}).get("total_score"),
        "export_preserved_rows": export_rows == expected_students,
        "no_hard_violations": not dashboard.get("score", {}).get("hard_violations"),
        "class_sizes_balanced_28_29": bool(class_sizes) and min(class_sizes) >= 28 and max(class_sizes) <= 29,
    }


def _exported_student_count(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        from openpyxl import load_workbook
    except ImportError:
        with zipfile.ZipFile(path) as archive:
            return 1 if "xl/workbook.xml" in archive.namelist() else 0
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return max(0, int(workbook.active.max_row or 0) - 1)
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
