from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT_PATH = Path("examples/generated_verified_200_students.xlsx")
CLASS_NOTE = "מומלץ ליצור פרויקט עם 8 כיתות: ז1, ז2, ז3, ז4, ז5, ז6, ז7, ז8."
CLASS_NAMES = [f"ז{index}" for index in range(1, 9)]

HEADERS = [
    "שם מלא",
    "מין",
    "בית ספר קודם",
    "ממוצע",
    "מתמטיקה",
    "אנגלית",
    "עברית",
    "התנהגות",
    "דומיננטיות",
    "חבר 1",
    "חבר 2",
    "חבר 3",
    "חייב להיות עם",
    "כיתות מותרות",
    "כיתות אסורות",
    "הערות מורה",
]

BOY_FIRST = [
    "אורי",
    "יואב",
    "דניאל",
    "איתמר",
    "ליאל",
    "עידו",
    "רועי",
    "עומר",
    "אריאל",
    "תומר",
    "גיא",
    "הלל",
    "איתן",
    "נדב",
    "רון",
    "שחר",
    "עמית",
    "אדם",
    "נועם",
    "יהונתן",
]
GIRL_FIRST = [
    "נועה",
    "מאיה",
    "תמר",
    "שירה",
    "רוני",
    "ליה",
    "יעל",
    "אלה",
    "מיקה",
    "איילה",
    "אביגיל",
    "מעיין",
    "טליה",
    "הילה",
    "אדוה",
    "עדי",
    "יובל",
    "שחר",
    "עמית",
    "אורי",
]
LAST_NAMES = [
    "כהן",
    "לוי",
    "מזרחי",
    "פרץ",
    "שלום",
    "חדד",
    "אוחנה",
    "בן דוד",
    "שגב",
    "ברק",
    "ישראלי",
    "אלמוג",
    "רוזן",
    "ביטון",
    "גולן",
    "אברהם",
    "סעדון",
    "קפלן",
    "מור",
    "דיין",
]
MIDDLE_NAMES = [
    "כרמל",
    "גליל",
    "תבור",
    "נגב",
    "ירדן",
    "מדבר",
    "כנרת",
    "שומרון",
    "ערבה",
    "בשן",
    "איילון",
    "יהודה",
    "גולן",
    "רמות",
    "אפק",
    "אורן",
    "תאנה",
    "שיטה",
    "רותם",
    "אלמוג",
]
SCHOOLS = [
    "אלון",
    "ברוש",
    "ארז",
    "דקל",
    "הדר",
    "ניצנים",
    "רימון",
    "שקד",
    "כרמים",
    "יובל",
]
BEHAVIOR = ["מצוין", "טובה", "טובה", "בינוני", "בינוני"]


def unique_name(index: int, gender: str, used_names: set[str]) -> str:
    first_pool = BOY_FIRST if gender == "בן" else GIRL_FIRST
    for attempt in range(len(first_pool) * len(LAST_NAMES)):
        value = index + attempt
        first = first_pool[value % len(first_pool)]
        middle = MIDDLE_NAMES[(index * 11 + attempt * 5) % len(MIDDLE_NAMES)]
        last = LAST_NAMES[((value // len(first_pool)) + (attempt * 3) + (5 if gender == "בת" else 0)) % len(LAST_NAMES)]
        name = f"{first} {middle} {last}"
        if name not in used_names:
            used_names.add(name)
            return name
    raise ValueError("לא ניתן ליצור שם ייחודי.")


def make_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    used_names: set[str] = set()
    group_size = 5
    group_count = 40
    class_source_slots: list[list[str]] = []
    for class_index in range(len(CLASS_NAMES)):
        extra_sources = {(class_index * 5 + offset) % len(SCHOOLS) for offset in range(5)}
        slots: list[str] = []
        for source_index, school in enumerate(SCHOOLS):
            slots.extend([school] * (3 if source_index in extra_sources else 2))
        class_source_slots.append(slots)

    for group_index in range(group_count):
        target_class_index = group_index % len(CLASS_NAMES)
        group_slot_in_class = group_index // len(CLASS_NAMES)
        gender_pattern = (
            ["בן", "בת", "בן", "בת", "בן"]
            if group_slot_in_class in {0, 2, 4}
            else ["בת", "בן", "בת", "בן", "בת"]
        )
        group_rows: list[dict[str, object]] = []
        for member_index in range(group_size):
            index = (group_index * group_size) + member_index
            class_slot = (group_slot_in_class * group_size) + member_index
            gender = gender_pattern[member_index]
            math = 62 + ((class_slot * 17) % 38)
            english = 61 + ((class_slot * 13) % 39)
            hebrew = 63 + ((class_slot * 11) % 37)
            average = round((math + english + hebrew) / 3)
            group_rows.append(
                {
                    "שם מלא": unique_name(index, gender, used_names),
                    "מין": gender,
                    "בית ספר קודם": class_source_slots[target_class_index][class_slot],
                    "ממוצע": average,
                    "מתמטיקה": math,
                    "אנגלית": english,
                    "עברית": hebrew,
                    "התנהגות": BEHAVIOR[class_slot % len(BEHAVIOR)],
                    "דומיננטיות": 1 + (class_slot % 5),
                    "חבר 1": "",
                    "חבר 2": "",
                    "חבר 3": "",
                    "חייב להיות עם": "",
                    "כיתות מותרות": CLASS_NAMES[target_class_index],
                    "כיתות אסורות": "",
                    "הערות מורה": "",
                }
            )
        names = [str(row["שם מלא"]) for row in group_rows]
        for member_index, row in enumerate(group_rows):
            friend_indices = [
                (member_index + 1) % group_size,
                (member_index + 2) % group_size,
                (member_index + 3) % group_size,
            ]
            row["חבר 1"] = names[friend_indices[0]]
            row["חבר 2"] = names[friend_indices[1]]
            row["חבר 3"] = names[friend_indices[2]]
            row["חייב להיות עם"] = names[(member_index + 1) % group_size]
        rows.extend(group_rows)
    return rows


def save(rows: list[dict[str, object]]) -> Path:
    OUT_PATH.parent.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])
    ws.auto_filter.ref = ws.dimensions
    ws["A1"].comment = Comment(
        "קובץ בדיקה תקין: 200 תלמידים, לכל תלמיד 3 חברים קיימים, ציוני עברית/מתמטיקה/אנגלית וממוצע. "
        + CLASS_NOTE,
        "Mosaicly",
    )

    fill = PatternFill("solid", fgColor="D9EAF7")
    side = Side(style="thin", color="D9E2EC")
    border = Border(left=side, right=side, top=side, bottom=side)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 30)

    wb.save(OUT_PATH)
    return OUT_PATH


def main() -> None:
    print(save(make_rows()).resolve())


if __name__ == "__main__":
    main()
