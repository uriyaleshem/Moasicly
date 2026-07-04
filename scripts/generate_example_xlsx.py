from __future__ import annotations

from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT_DIR = Path("examples")
OUT_DIR.mkdir(exist_ok=True)
random.seed(42621)

CLASS_NAMES = ["ז׳1", "ז׳2", "ז׳3", "ז׳4", "ז׳5", "ז׳6"]
SCHOOLS = ["אלון", "ברוש", "ארז", "דקל", "הדר", "ניצנים", "רימון", "שקד"]
BOY_FIRST = ["אורי", "נועם", "יואב", "אדם", "דניאל", "איתמר", "ליאל", "עידו", "רועי", "עומר", "יהונתן", "אריאל", "תומר", "גיא", "הלל", "איתן", "נדב", "רון", "שחר", "עמית"]
GIRL_FIRST = ["נועה", "מאיה", "תמר", "שירה", "רוני", "ליה", "יעל", "אורי", "אלה", "מיקה", "איילה", "אביגיל", "מעיין", "טליה", "הילה", "אדוה", "עדי", "יובל", "שחר", "עמית"]
LAST_NAMES = ["כהן", "לוי", "מזרחי", "פרץ", "שלום", "חדד", "אוחנה", "בן דוד", "שגב", "ברק", "ישראלי", "אלמוג", "רוזן", "ביטון", "גולן", "אברהם", "סעדון", "קפלן", "מור", "דיין", "שטרן", "הררי", "צור", "עמר", "טל"]
BEHAVIOR = ["מצוין", "טובה", "טובה", "בינוני", "בינוני", "בעייתית"]
NOTES = ["", "זקוק/ה לחבר מוכר", "מנהיג/ה חיובי/ת", "כדאי להפריד ממוקד רעש", "מתאים/ה למסגרת שקטה", "חזק/ה חברתית", "רגיש/ה במעברים"]

CLEAN_HEADERS = [
    "שם מלא",
    "מין",
    "בית ספר קודם",
    "ממוצע",
    "מתמטיקה",
    "אנגלית",
    "עברית",
    "התנהגות",
    "דומיננטיות",
    "חבר 1",
    "חבר 2",
    "חבר 3",
    "כיתות מותרות",
    "כיתות אסורות",
    "חייב להיות עם",
    "אסור להיות עם",
    "הערות הורים",
    "הערות מורה",
    "הערות ראיון",
]

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def unique_name(index: int, gender: str) -> str:
    first_pool = BOY_FIRST if gender == "בן" else GIRL_FIRST
    local_index = index // 2
    first = first_pool[local_index % len(first_pool)]
    last = LAST_NAMES[(local_index * 7 + (0 if gender == "בן" else 3)) % len(LAST_NAMES)]
    suffix = local_index // (len(first_pool) * len(LAST_NAMES))
    return f"{first} {last}" + (f" {suffix + 1}" if suffix else "")


def make_rows(count: int, classes: list[str], *, messy_values: bool = False, sparse: bool = False) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for index in range(count):
        gender = "בן" if index % 2 == 0 else "בת"
        base = 66 + ((index * 11) % 31)
        math = max(45, min(100, base + ((index % 9) - 4)))
        english = max(45, min(100, base + (((index * 2) % 11) - 5)))
        hebrew = max(45, min(100, base + (((index * 3) % 9) - 4)))
        row: dict[str, object] = {
            "שם מלא": unique_name(index, gender),
            "מין": gender,
            "בית ספר קודם": SCHOOLS[(index * 3 + index // 7) % len(SCHOOLS)],
            "ממוצע": round((math + english + hebrew) / 3),
            "מתמטיקה": math,
            "אנגלית": english,
            "עברית": hebrew,
            "התנהגות": BEHAVIOR[(index * 5) % len(BEHAVIOR)],
            "דומיננטיות": 1 + ((index * 3) % 5),
            "חבר 1": "",
            "חבר 2": "",
            "חבר 3": "",
            "כיתות מותרות": "",
            "כיתות אסורות": "",
            "חייב להיות עם": "",
            "אסור להיות עם": "",
            "הערות הורים": "",
            "הערות מורה": NOTES[(index * 4) % len(NOTES)],
            "הערות ראיון": "",
        }
        if sparse and index % 9 == 0:
            row["ממוצע"] = ""
        if sparse and index % 11 == 0:
            row["מין"] = ""
        if messy_values:
            if row["מין"] == "בן":
                row["מין"] = "זכר" if index % 4 else "M"
            elif row["מין"] == "בת":
                row["מין"] = "נקבה" if index % 5 else "F"
            if row["התנהגות"] == "מצוין":
                row["התנהגות"] = "מצויין"
            elif row["התנהגות"] == "טובה":
                row["התנהגות"] = "טוב"
        rows.append(row)

    by_school: dict[str, list[int]] = {}
    for index, row in enumerate(rows):
        by_school.setdefault(str(row["בית ספר קודם"]), []).append(index)

    for index, row in enumerate(rows):
        same_school = by_school[str(row["בית ספר קודם"])]
        if len(same_school) > 1:
            friend_index = same_school[(same_school.index(index) + 1) % len(same_school)]
            row["חבר 1"] = rows[friend_index]["שם מלא"]
        if index % 5 == 0 and count > 3:
            row["חבר 2"] = rows[(index + 3) % count]["שם מלא"]
        if index % 13 == 0 and count > 10:
            row["חבר 3"] = rows[(index + 10) % count]["שם מלא"]
        if index % 17 == 0:
            row["כיתות אסורות"] = classes[(index // 17) % len(classes)]
        if index % 23 == 0:
            row["כיתות מותרות"] = "|".join(classes[: min(3, len(classes))])
        if index % 29 == 0 and index + 1 < count:
            row["חייב להיות עם"] = rows[index + 1]["שם מלא"]
        if index % 31 == 0 and index + 2 < count:
            row["אסור להיות עם"] = rows[index + 2]["שם מלא"]
    return rows


def rows_to_matrix(rows: list[dict[str, object]], headers: list[str]) -> list[list[object]]:
    return [[row.get(header, "") for header in headers] for row in rows]


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]], note: str | None = None) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if note:
        ws["A1"].comment = Comment(note, "Mosaicly")

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 28)


def save_workbook(path: Path, sheets: list[tuple[str, list[str], list[list[object]], str | None]]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows, note in sheets:
        add_sheet(wb, title, headers, rows, note)
    wb.save(path)
    return path


def create_clean() -> Path:
    rows = make_rows(48, CLASS_NAMES[:4])
    return save_workbook(
        OUT_DIR / "generated_clean_48_students.xlsx",
        [("תלמידים", CLEAN_HEADERS, rows_to_matrix(rows, CLEAN_HEADERS), "תרחיש נקי יחסית: 48 תלמידים, 4 כיתות, חברים ואילוצים מתונים.")],
    )


def create_easy_48() -> Path:
    rows = make_rows(48, CLASS_NAMES[:4])
    for row in rows:
        row["חבר 1"] = ""
        row["חבר 2"] = ""
        row["חבר 3"] = ""
        row["כיתות מותרות"] = ""
        row["כיתות אסורות"] = ""
        row["חייב להיות עם"] = ""
        row["אסור להיות עם"] = ""
    for index in range(0, len(rows), 2):
        if index + 1 >= len(rows):
            break
        rows[index]["חבר 1"] = rows[index + 1]["שם מלא"]
        rows[index + 1]["חבר 1"] = rows[index]["שם מלא"]
    for index in range(0, len(rows), 8):
        if index + 3 < len(rows):
            rows[index]["חבר 2"] = rows[index + 3]["שם מלא"]
            rows[index + 3]["חבר 2"] = rows[index]["שם מלא"]
    return save_workbook(
        OUT_DIR / "generated_easy_48_students.xlsx",
        [("תלמידים", CLEAN_HEADERS, rows_to_matrix(rows, CLEAN_HEADERS), "תרחיש קל יותר: 48 תלמידים, 4 כיתות, בקשות חברים זוגיות וללא אילוצי כיתה קשיחים.")],
    )


def create_large() -> Path:
    rows = make_rows(180, CLASS_NAMES[:6])
    return save_workbook(
        OUT_DIR / "generated_large_180_students.xlsx",
        [("תלמידים", CLEAN_HEADERS, rows_to_matrix(rows, CLEAN_HEADERS), "תרחיש גדול יותר לבדיקת ביצועים: 180 תלמידים, 6 כיתות.")],
    )


def create_messy() -> Path:
    source_rows = make_rows(64, CLASS_NAMES[:4], messy_values=True, sparse=True)
    messy_headers = [
        "תלמיד/ה",
        "בן/בת",
        "יסודי",
        "grade avg",
        "math",
        "english grade",
        "שפה",
        "תפקוד",
        "מנהיגות",
        "בחירת חבר 1",
        "friend2",
        "friend 3",
        "allowed",
        "forbidden",
        "together",
        "separate",
        "parents notes",
        "notes",
        "interview",
    ]
    source_to_messy = dict(zip(CLEAN_HEADERS, messy_headers, strict=True))
    messy_rows = [{source_to_messy[header]: row.get(header, "") for header in CLEAN_HEADERS} for row in source_rows]
    return save_workbook(
        OUT_DIR / "generated_messy_headers_values.xlsx",
        [("יבוא מבולגן", messy_headers, rows_to_matrix(messy_rows, messy_headers), "בודק מיפוי אוטומטי ונרמול ערכים: כותרות מעורבות, זכר/נקבה/M/F, וחוסרים חלקיים.")],
    )


def create_conflicts() -> Path:
    rows = make_rows(36, CLASS_NAMES[:3])
    rows[0]["כיתות מותרות"] = "ז׳1"
    rows[0]["כיתות אסורות"] = "ז׳1"
    rows[0]["הערות מורה"] = "אילוץ סותר בכוונה: מותר ואסור באותה כיתה."
    rows[1]["חייב להיות עם"] = rows[2]["שם מלא"]
    rows[1]["אסור להיות עם"] = rows[2]["שם מלא"]
    rows[1]["הערות מורה"] = "זוג שהוגדר גם יחד וגם בנפרד."
    rows[3]["חבר 1"] = rows[3]["שם מלא"]
    rows[3]["הערות מורה"] = "בקשת חבר לעצמו לבדיקה."
    rows[4]["חבר 1"] = "שם שלא קיים בקובץ"
    rows[4]["הערות מורה"] = "חבר שלא קיים בקובץ."
    for index in range(5, 18):
        rows[index]["כיתות מותרות"] = "ז׳1"
    return save_workbook(
        OUT_DIR / "generated_constraints_conflicts.xlsx",
        [("אילוצים לבדיקה", CLEAN_HEADERS, rows_to_matrix(rows, CLEAN_HEADERS), "כולל בכוונה אילוצים בעייתיים כדי לבדוק מסכי בדיקה/אילוצים מתנגשים.")],
    )


def create_multisheet() -> Path:
    clean_rows = make_rows(40, CLASS_NAMES[:4])
    sparse_rows = make_rows(32, CLASS_NAMES[:4], messy_values=True, sparse=True)
    info_headers = ["שם גיליון", "מה לבדוק"]
    info_rows = [
        ["תלמידים", "גיליון ראשון נקי יחסית ומתאים לייבוא רגיל."],
        ["נתונים חסרים", "אותם שדות עם חלק מהציונים/מגדרים חסרים וערכים לנרמול."],
    ]
    return save_workbook(
        OUT_DIR / "generated_multisheet_students.xlsx",
        [
            ("תלמידים", CLEAN_HEADERS, rows_to_matrix(clean_rows, CLEAN_HEADERS), "גיליון ראשי לייבוא."),
            ("נתונים חסרים", CLEAN_HEADERS, rows_to_matrix(sparse_rows, CLEAN_HEADERS), "בחרו בגיליון הזה כדי לבדוק החלפת Sheet ותיקוף חוסרים."),
            ("הסבר", info_headers, info_rows, "גיליון הסבר, לא מיועד לייבוא תלמידים."),
        ],
    )


def main() -> None:
    created = [create_clean(), create_easy_48(), create_large(), create_messy(), create_conflicts(), create_multisheet()]
    for path in created:
        print(path.resolve())


if __name__ == "__main__":
    main()
