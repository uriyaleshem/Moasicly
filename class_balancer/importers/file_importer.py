from __future__ import annotations

import csv
import importlib.util
import re
import zipfile
from dataclasses import dataclass, field
from itertools import islice
from pathlib import Path
from typing import Any

try:
    from defusedxml import ElementTree as ET
except ImportError:  # pragma: no cover - dependency is declared, fallback keeps basic installs working.
    from xml.etree import ElementTree as ET


MAX_IMPORT_FILE_BYTES = 25 * 1024 * 1024
MAX_IMPORT_ROWS = 100_000
MAX_IMPORT_COLUMNS = 500
MAX_CELL_CHARS = 8_000
MAX_ZIP_ENTRIES = 300
MAX_ZIP_UNCOMPRESSED_BYTES = 80 * 1024 * 1024
MAX_XML_BYTES = 15 * 1024 * 1024
MAX_ZIP_COMPRESSION_RATIO = 100


@dataclass(slots=True)
class ImportedTable:
    path: Path
    headers: list[str]
    rows: list[dict[str, Any]]
    row_count: int
    sheet_names: list[str] = field(default_factory=list)
    selected_sheet: str = ""
    encoding: str = ""

    def preview(self, limit: int = 20) -> list[dict[str, Any]]:
        return self.rows[:limit]


def load_table(path: str | Path, sheet_name: str | None = None, preview_limit: int | None = None) -> ImportedTable:
    file_path = Path(path)
    _validate_file_size(file_path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(file_path, preview_limit=preview_limit)
    if suffix == ".xlsx":
        if importlib.util.find_spec("openpyxl"):
            return _load_xlsx_openpyxl(file_path, sheet_name=sheet_name, preview_limit=preview_limit)
        return _load_xlsx_stdlib(file_path, sheet_name=sheet_name, preview_limit=preview_limit)
    raise ValueError("סוג קובץ לא נתמך. יש לבחור CSV או XLSX.")


def _load_csv(path: Path, preview_limit: int | None = None) -> ImportedTable:
    encoding = _detect_encoding(path)
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        headers_raw = next(reader, [])

        if not headers_raw:
            return ImportedTable(path=path, headers=[], rows=[], row_count=0, encoding=encoding)

        headers = _dedupe_headers([_clean_cell(value) for value in headers_raw[:MAX_IMPORT_COLUMNS]])
        rows: list[dict[str, Any]] = []
        count = 0
        for raw in reader:
            _validate_row_shape(raw)
            if not any(_clean_cell(value) for value in raw):
                continue
            count += 1
            if count > MAX_IMPORT_ROWS:
                raise ValueError(f"קובץ CSV גדול מדי. המגבלה היא {MAX_IMPORT_ROWS:,} שורות נתונים.")
            if preview_limit is None or len(rows) < preview_limit:
                rows.append({header: _clean_cell(raw[index]) if index < len(raw) else "" for index, header in enumerate(headers)})

    if not headers:
        return ImportedTable(path=path, headers=[], rows=[], row_count=0, encoding=encoding)
    return ImportedTable(path=path, headers=headers, rows=rows, row_count=count, encoding=encoding)


def _detect_encoding(path: Path) -> str:
    data = path.read_bytes()[:8192]
    for encoding in ("utf-8-sig", "utf-8", "cp1255", "windows-1255", "cp1252", "latin-1"):
        try:
            data.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8-sig"


def _validate_file_size(path: Path) -> None:
    try:
        size = path.stat().st_size
    except OSError:
        return
    if size > MAX_IMPORT_FILE_BYTES:
        raise ValueError(f"הקובץ גדול מדי לייבוא. המגבלה היא {MAX_IMPORT_FILE_BYTES // (1024 * 1024)}MB.")


def _validate_row_shape(values: Any) -> None:
    if len(values) > MAX_IMPORT_COLUMNS:
        raise ValueError(f"הקובץ מכיל יותר מדי עמודות. המגבלה היא {MAX_IMPORT_COLUMNS:,}.")
    for value in values:
        if len(_clean_cell(value)) > MAX_CELL_CHARS:
            raise ValueError(f"תא בקובץ ארוך מדי. המגבלה היא {MAX_CELL_CHARS:,} תווים לתא.")


def _count_csv_data_rows(path: Path, encoding: str) -> int:
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle)
        return sum(1 for row in islice(reader, 1, MAX_IMPORT_ROWS + 1) if any(_clean_cell(value) for value in row))


def _load_xlsx_openpyxl(path: Path, sheet_name: str | None = None, preview_limit: int | None = None) -> ImportedTable:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[selected]
        if sheet.max_column and sheet.max_column > MAX_IMPORT_COLUMNS:
            raise ValueError(f"קובץ XLSX מכיל יותר מדי עמודות. המגבלה היא {MAX_IMPORT_COLUMNS:,}.")
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = _dedupe_headers([_clean_cell(value) for value in next(iterator)[:MAX_IMPORT_COLUMNS]])
        except StopIteration:
            return ImportedTable(path=path, headers=[], rows=[], row_count=0, sheet_names=workbook.sheetnames, selected_sheet=selected)

        rows: list[dict[str, Any]] = []
        count = 0
        for values in iterator:
            _validate_row_shape(values)
            if not any(_clean_cell(value) for value in values):
                continue
            count += 1
            if count > MAX_IMPORT_ROWS:
                raise ValueError(f"קובץ XLSX גדול מדי. המגבלה היא {MAX_IMPORT_ROWS:,} שורות נתונים.")
            if preview_limit is None or len(rows) < preview_limit:
                rows.append({header: _clean_cell(values[index]) if index < len(values) else "" for index, header in enumerate(headers)})

        return ImportedTable(
            path=path,
            headers=headers,
            rows=rows,
            row_count=count,
            sheet_names=workbook.sheetnames,
            selected_sheet=selected,
        )
    finally:
        workbook.close()


def _load_xlsx_stdlib(path: Path, sheet_name: str | None = None, preview_limit: int | None = None) -> ImportedTable:
    with zipfile.ZipFile(path) as archive:
        _validate_zip_archive(archive)
        sheets = _read_workbook_sheets(archive)
        if not sheets:
            return ImportedTable(path=path, headers=[], rows=[], row_count=0)
        selected_name = sheet_name if sheet_name in sheets else next(iter(sheets))
        shared_strings = _read_shared_strings(archive)
        matrix = _read_sheet_matrix(archive, sheets[selected_name], shared_strings)

    while matrix and not any(_clean_cell(value) for value in matrix[0]):
        matrix.pop(0)
    if not matrix:
        return ImportedTable(path=path, headers=[], rows=[], row_count=0, sheet_names=list(sheets), selected_sheet=selected_name)

    headers = _dedupe_headers([_clean_cell(value) for value in matrix[0]])
    rows: list[dict[str, Any]] = []
    count = 0
    for raw in matrix[1:]:
        if not any(_clean_cell(value) for value in raw):
            continue
        count += 1
        if preview_limit is None or len(rows) < preview_limit:
            rows.append({header: _clean_cell(raw[index]) if index < len(raw) else "" for index, header in enumerate(headers)})

    return ImportedTable(
        path=path,
        headers=headers,
        rows=rows,
        row_count=count,
        sheet_names=list(sheets),
        selected_sheet=selected_name,
    )


def _validate_zip_archive(archive: zipfile.ZipFile) -> None:
    entries = archive.infolist()
    if len(entries) > MAX_ZIP_ENTRIES:
        raise ValueError(f"קובץ XLSX מכיל יותר מדי רכיבי ZIP. המגבלה היא {MAX_ZIP_ENTRIES:,}.")
    total_uncompressed = 0
    for info in entries:
        total_uncompressed += int(info.file_size)
        if info.file_size > MAX_XML_BYTES and info.filename.endswith(".xml"):
            raise ValueError("קובץ XLSX מכיל XML גדול מדי.")
        if info.compress_size and info.file_size / max(1, info.compress_size) > MAX_ZIP_COMPRESSION_RATIO:
            raise ValueError("קובץ XLSX נדחה כי יחס הדחיסה חשוד.")
    if total_uncompressed > MAX_ZIP_UNCOMPRESSED_BYTES:
        raise ValueError("קובץ XLSX גדול מדי לאחר פתיחת ZIP.")


def _read_zip_member(archive: zipfile.ZipFile, name: str, max_bytes: int) -> bytes:
    try:
        info = archive.getinfo(name)
    except KeyError as exc:
        raise ValueError(f"קובץ XLSX חסר רכיב נדרש: {name}") from exc
    if info.file_size > max_bytes:
        raise ValueError(f"רכיב XLSX גדול מדי: {name}")
    data = archive.read(name)
    if len(data) > max_bytes:
        raise ValueError(f"רכיב XLSX גדול מדי: {name}")
    return data


def _read_workbook_sheets(archive: zipfile.ZipFile) -> dict[str, str]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    workbook = ET.fromstring(_read_zip_member(archive, "xl/workbook.xml", MAX_XML_BYTES))
    rels = ET.fromstring(_read_zip_member(archive, "xl/_rels/workbook.xml.rels", MAX_XML_BYTES))
    targets: dict[str, str] = {}
    for rel in rels.findall("pkgrel:Relationship", ns):
        target = rel.attrib.get("Target", "")
        if target.startswith("/"):
            target = target.lstrip("/")
        elif not target.startswith("xl/"):
            target = "xl/" + target
        targets[rel.attrib["Id"]] = target

    sheets: dict[str, str] = {}
    for sheet in workbook.findall(".//main:sheet", ns):
        name = sheet.attrib.get("name", f"Sheet {len(sheets) + 1}")
        rel_id = sheet.attrib.get(f"{{{ns['rel']}}}id", "")
        target = targets.get(rel_id)
        if target:
            sheets[name] = target
    return sheets


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(_read_zip_member(archive, "xl/sharedStrings.xml", MAX_XML_BYTES))
    values: list[str] = []
    for item in root.findall("main:si", ns):
        if len(values) >= MAX_IMPORT_ROWS * MAX_IMPORT_COLUMNS:
            raise ValueError("קובץ XLSX מכיל יותר מדי shared strings.")
        texts = [node.text or "" for node in item.findall(".//main:t", ns)]
        values.append("".join(texts))
    return values


def _read_sheet_matrix(archive: zipfile.ZipFile, sheet_path: str, shared_strings: list[str]) -> list[list[Any]]:
    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(_read_zip_member(archive, sheet_path, MAX_XML_BYTES))
    rows: list[list[Any]] = []
    for row in root.findall(".//main:row", ns):
        if len(rows) >= MAX_IMPORT_ROWS + 1:
            raise ValueError(f"קובץ XLSX גדול מדי. המגבלה היא {MAX_IMPORT_ROWS:,} שורות נתונים.")
        values: dict[int, Any] = {}
        max_index = 0
        for cell in row.findall("main:c", ns):
            ref = cell.attrib.get("r", "")
            index = _column_index(ref)
            if index > MAX_IMPORT_COLUMNS:
                raise ValueError(f"קובץ XLSX מכיל יותר מדי עמודות. המגבלה היא {MAX_IMPORT_COLUMNS:,}.")
            max_index = max(max_index, index)
            values[index] = _xlsx_cell_value(cell, shared_strings, ns)
        rows.append([values.get(index, "") for index in range(1, max_index + 1)])
    return rows


def _xlsx_cell_value(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> Any:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.findall(".//main:t", ns)]
        return "".join(texts)
    value_node = cell.find("main:v", ns)
    if value_node is None:
        return ""
    raw = value_node.text or ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if cell_type == "b":
        return "TRUE" if raw == "1" else "FALSE"
    return raw


def _column_index(reference: str) -> int:
    if not reference:
        return 1
    match = re.fullmatch(r"([A-Za-z]{1,3})[1-9]\d{0,6}", reference)
    if not match:
        raise ValueError(f"הפניה לא תקינה לתא XLSX: {reference[:40]}")
    letters = match.group(1).upper()
    total = 0
    for char in letters:
        total = total * 26 + (ord(char) - ord("A") + 1)
    if total > MAX_IMPORT_COLUMNS:
        raise ValueError(f"קובץ XLSX מכיל יותר מדי עמודות. המגבלה היא {MAX_IMPORT_COLUMNS:,}.")
    return total or 1


def _dedupe_headers(headers: list[str]) -> list[str]:
    deduped: list[str] = []
    counts: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        base = header or f"עמודה {index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        deduped.append(base if count == 1 else f"{base} ({count})")
    return deduped


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()
